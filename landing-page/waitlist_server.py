#!/usr/bin/env python3
"""
Placedon — local development server.

Handles:
  /register  POST  → create account, session cookie, redirect to dashboard
  /login     POST  → verify credentials, session cookie, redirect to dashboard
  /logout    GET   → clear session, redirect to home
  /submit    POST  → waitlist form → waitlist.xlsx
  All other GET    → static files from landing-page/

Run:
  python3 waitlist_server.py

Data files created automatically in ./data/
  waitlist.xlsx  — waitlist signups
  users.json     — registered accounts (passwords hashed with pbkdf2)
"""

import os, json, csv, secrets, hashlib
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import parse_qs, unquote_plus
from datetime import datetime

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

PORT     = 8765
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')

XLSX_PATH  = os.path.join(DATA_DIR, 'waitlist.xlsx')
CSV_PATH   = os.path.join(DATA_DIR, 'waitlist.csv')
USERS_PATH = os.path.join(DATA_DIR, 'users.json')

WAITLIST_COLS   = ['Timestamp','Name','Email','Company / Firm','Role','Top compliance question','Source']
WAITLIST_WIDTHS = [22, 24, 34, 26, 34, 55, 16]

MIME = {
    '.html': 'text/html; charset=utf-8',
    '.css':  'text/css',
    '.js':   'application/javascript',
    '.ttf':  'font/ttf',
    '.woff2':'font/woff2',
    '.png':  'image/png',
    '.jpg':  'image/jpeg',
    '.svg':  'image/svg+xml',
    '.ico':  'image/x-icon',
}

# ── In-memory session store ────────────────────────────────────────────────────
# session_id → user dict  (survives the process, lost on restart — fine for dev)
_sessions: dict[str, dict] = {}


# ── Password helpers ───────────────────────────────────────────────────────────

def _hash_password(password: str) -> str:
    salt = os.urandom(16)
    dk   = hashlib.pbkdf2_hmac('sha256', password.encode(), salt, 260_000)
    return salt.hex() + ':' + dk.hex()

def _verify_password(password: str, stored: str) -> bool:
    try:
        salt_hex, dk_hex = stored.split(':', 1)
        salt = bytes.fromhex(salt_hex)
        dk   = hashlib.pbkdf2_hmac('sha256', password.encode(), salt, 260_000)
        return secrets.compare_digest(dk.hex(), dk_hex)
    except Exception:
        return False


# ── User store (users.json) ────────────────────────────────────────────────────

def _load_users() -> list:
    if not os.path.exists(USERS_PATH):
        return []
    with open(USERS_PATH, encoding='utf-8') as f:
        return json.load(f)

def _save_users(users: list):
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(USERS_PATH, 'w', encoding='utf-8') as f:
        json.dump(users, f, indent=2, ensure_ascii=False)

def _find_user(email: str):
    for u in _load_users():
        if u['email'].lower() == email.lower():
            return u
    return None

def _create_user(name: str, email: str, password: str, role: str) -> dict:
    users = _load_users()
    user  = {
        'name':          name,
        'email':         email.lower(),
        'password_hash': _hash_password(password),
        'role':          role,
        'created_at':    datetime.now().isoformat(timespec='seconds'),
    }
    users.append(user)
    _save_users(users)
    return user


# ── Session helpers ────────────────────────────────────────────────────────────

def _new_session(user: dict) -> str:
    sid = secrets.token_urlsafe(40)
    _sessions[sid] = {
        'name':       user['name'],
        'email':      user['email'],
        'role':       user.get('role', ''),
        'created_at': user.get('created_at', ''),
    }
    return sid

def _get_session(headers):
    raw = headers.get('Cookie', '')
    for part in raw.split(';'):
        part = part.strip()
        if part.startswith('placedon_sid='):
            return _sessions.get(part[len('placedon_sid='):])
    return None

def _clear_session(headers):
    raw = headers.get('Cookie', '')
    for part in raw.split(';'):
        part = part.strip()
        if part.startswith('placedon_sid='):
            sid = part[len('placedon_sid='):]
            _sessions.pop(sid, None)
            return sid
    return None


# ── Waitlist Excel helpers ─────────────────────────────────────────────────────

def _init_xlsx():
    os.makedirs(DATA_DIR, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Waitlist'
    ws.append(WAITLIST_COLS)
    for i, cell in enumerate(ws[1], 1):
        cell.font      = Font(bold=True, color='F5F3EF', name='Calibri')
        cell.fill      = PatternFill(fill_type='solid', fgColor='334155')
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.column_dimensions[cell.column_letter].width = WAITLIST_WIDTHS[i - 1]
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'
    wb.save(XLSX_PATH)

def _append_xlsx(row: list):
    if not os.path.exists(XLSX_PATH):
        _init_xlsx()
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    ws.append(row)
    idx = ws.max_row
    if idx % 2 == 0:
        fill = PatternFill(fill_type='solid', fgColor='EFEDE9')
        for cell in ws[idx]:
            cell.fill = fill
    wb.save(XLSX_PATH)

def _append_csv(row: list):
    os.makedirs(DATA_DIR, exist_ok=True)
    need_header = not os.path.exists(CSV_PATH)
    with open(CSV_PATH, 'a', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        if need_header:
            w.writerow(WAITLIST_COLS)
        w.writerow(row)

def _save_waitlist(data: dict):
    row = [
        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        data.get('name', '').strip(),
        data.get('email', '').strip(),
        data.get('company', '').strip(),
        data.get('role', '').strip(),
        data.get('question', '').strip(),
        data.get('source', 'landing-page').strip(),
    ]
    if HAS_EXCEL:
        _append_xlsx(row)
        print(f"  [waitlist] {row[2]}  |  {row[4]}")
    else:
        _append_csv(row)
        print(f"  [waitlist] {row[2]}  |  {row[4]}  (CSV)")


# ── HTML helpers ───────────────────────────────────────────────────────────────

def _inject_user(html: bytes, user: dict) -> bytes:
    """Inject user data into dashboard HTML as a JS variable."""
    script = (
        '<script>window.__PLACEDON_USER__='
        + json.dumps(user, ensure_ascii=False)
        + ';</script>'
    )
    return html.replace(b'</head>', script.encode() + b'</head>', 1)


# ── HTTP handler ───────────────────────────────────────────────────────────────

class Handler(BaseHTTPRequestHandler):

    def log_message(self, fmt, *args):
        print(f"[{datetime.now().strftime('%H:%M:%S')}]  {fmt % args}")

    # ── GET ──────────────────────────────────────────────────────────────────

    def do_GET(self):
        path = self.path.split('?')[0].rstrip('/')

        # Canonical redirects
        if path in ('', '/'):
            return self._redirect('/index.html')
        if path == '/waitlist':
            return self._redirect('/waitlist.html')
        if path == '/auth':
            return self._redirect('/auth.html')
        if path == '/dashboard':
            return self._redirect('/dashboard.html')

        # Logout
        if path == '/logout':
            _clear_session(self.headers)
            self.send_response(302)
            self.send_header('Set-Cookie',
                'placedon_sid=; Path=/; HttpOnly; Max-Age=0')
            self.send_header('Location', '/index.html')
            self.end_headers()
            return

        # Dashboard — requires session
        if path == '/dashboard.html':
            user = _get_session(self.headers)
            if not user:
                return self._redirect('/auth.html?mode=signin')
            html = self._read_file('dashboard.html')
            if html is None:
                return self._respond(404, b'Not found')
            html = _inject_user(html, user)
            self.send_response(200)
            self.send_header('Content-Type', 'text/html; charset=utf-8')
            self.send_header('Content-Length', str(len(html)))
            self.end_headers()
            self.wfile.write(html)
            return

        # Static files
        rel   = path.lstrip('/')
        fpath = os.path.realpath(os.path.join(BASE_DIR, *rel.split('/')))
        if not fpath.startswith(BASE_DIR):
            return self._respond(403, b'Forbidden')
        if os.path.isfile(fpath):
            _, ext = os.path.splitext(fpath)
            mime   = MIME.get(ext.lower(), 'application/octet-stream')
            with open(fpath, 'rb') as f:
                body = f.read()
            self.send_response(200)
            self.send_header('Content-Type', mime)
            self.send_header('Content-Length', str(len(body)))
            self.end_headers()
            self.wfile.write(body)
        else:
            self._respond(404, b'Not found')

    # ── POST ─────────────────────────────────────────────────────────────────

    def do_POST(self):
        length = int(self.headers.get('Content-Length', 0))
        body   = self.rfile.read(length).decode('utf-8', errors='replace')
        raw    = parse_qs(body, keep_blank_values=True)
        data   = {k: unquote_plus(v[0]).strip() for k, v in raw.items()}

        if self.path == '/submit':
            self._handle_waitlist(data)
        elif self.path == '/register':
            self._handle_register(data)
        elif self.path == '/login':
            self._handle_login(data)
        else:
            self._respond(404, b'Not found')

    # ── Route handlers ────────────────────────────────────────────────────────

    def _handle_waitlist(self, data: dict):
        _save_waitlist(data)
        self._redirect('/thankyou.html')

    def _handle_register(self, data: dict):
        name     = data.get('name', '')
        email    = data.get('email', '')
        password = data.get('password', '')
        role     = data.get('role', '')

        if not all([name, email, password, role]):
            return self._redirect('/auth.html?mode=signup&error=missing_fields')
        if len(password) < 8:
            return self._redirect('/auth.html?mode=signup&error=weak_password')
        if _find_user(email):
            return self._redirect('/auth.html?mode=signup&error=email_exists')

        user = _create_user(name, email, password, role)
        sid  = _new_session(user)
        print(f"  [register] {email}  |  {role}")

        self.send_response(302)
        self.send_header('Set-Cookie',
            f'placedon_sid={sid}; Path=/; HttpOnly; Max-Age=604800')
        self.send_header('Location', '/dashboard.html')
        self.end_headers()

    def _handle_login(self, data: dict):
        email    = data.get('email', '')
        password = data.get('password', '')

        if not email or not password:
            return self._redirect('/auth.html?mode=signin&error=missing_fields')

        user = _find_user(email)
        if not user or not _verify_password(password, user['password_hash']):
            return self._redirect('/auth.html?mode=signin&error=invalid_credentials')

        sid = _new_session(user)
        print(f"  [login]    {email}")

        self.send_response(302)
        self.send_header('Set-Cookie',
            f'placedon_sid={sid}; Path=/; HttpOnly; Max-Age=604800')
        self.send_header('Location', '/dashboard.html')
        self.end_headers()

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _respond(self, code: int, body: bytes = b''):
        self.send_response(code)
        self.send_header('Content-Length', str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _redirect(self, location: str):
        self.send_response(302)
        self.send_header('Location', location)
        self.end_headers()

    def _read_file(self, filename: str):
        path = os.path.join(BASE_DIR, filename)
        if os.path.isfile(path):
            with open(path, 'rb') as f:
                return f.read()
        return None


# ── Entry point ────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    os.makedirs(DATA_DIR, exist_ok=True)
    if HAS_EXCEL and not os.path.exists(XLSX_PATH):
        _init_xlsx()

    store = 'data/waitlist.xlsx' if HAS_EXCEL else 'data/waitlist.csv'
    print()
    print('  Placedon — local server')
    print(f'  http://localhost:{PORT}')
    print(f'  Waitlist  → {store}')
    print(f'  Users     → data/users.json')
    print('  Press Ctrl+C to stop.')
    print()

    server = HTTPServer(('localhost', PORT), Handler)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print('\n  Server stopped.')
